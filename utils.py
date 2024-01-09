import os
import pickle
import re
from datetime import datetime, timedelta
from typing import List

import numpy as np
import numpy.typing as npt
import pandas as pd
import plotly.graph_objects as go
import requests
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from plotly.colors import n_colors
from pyhtml2pdf import converter
from pytz import timezone
from scipy.stats import norm

from variables import (current_results_path, default_headers, dict_index_stock,
                       get_default_values, list_emails, old_results_path,
                       result_css, src_css, src_html, temp_results_path)


#
def resolve_relative_path(path: str) -> str:
    """
    Helper function to resolve relative paths
    """
    return os.path.abspath(os.path.join(os.path.dirname(__file__), path))


def create_folder(path: str) -> None:
    """
    Create a foder in case it has not been created already
    """
    if not os.path.exists(path):
        os.mkdir(path)
        print(f"Directory created: {path}")


#  Auxiliary functions for styling the table

def scale_col_range(col: pd.Series, range: int) -> pd.Series:
    """Scale column to range
    col = [0,10,100], range = 10 => [0,1,10]
    """
    return ((range) * ((col - col.min()) / (col.max() - col.min() + 1e-9))).astype(np.int32)


def bold(col: pd.Series) -> List:
    """
    Return list string in col in bold using HTML syntax
    """
    return [f"<b>{element}</b>" for element in col]


def posneg_binary_color(col: pd.Series, pos_color, neg_color):
    """
    Return series with colors to differentiate positive and negative values
    """
    aux = col.reset_index()
    aux['color'] = neg_color
    aux.loc[col >= 0, 'color'] = pos_color
    return np.array(aux['color'])


def posneg_gradient(col: pd.Series) -> npt.NDArray:
    """
    Return series with colors to differentiate positive and negative values
    """
    aux = col.reset_index()
    aux['r'] = aux['g'] = aux['b'] = 255
    aux.loc[aux.Änderung < 0, 'g'] = aux.loc[aux.Änderung < 0,
                                             'b'] = 255 - 150 * aux.Änderung / (aux.Änderung.min())
    aux.loc[aux.Änderung > 0, 'g'] = aux.loc[aux.Änderung > 0,
                                             'r'] = 255 - 150 * aux.Änderung / (aux.Änderung.max())
    aux = aux.astype(np.int32)  # type: ignore
    return np.array([f"rgb({aux.loc[i,'r']},{aux.loc[i,'g']},{aux.loc[i,'b']})" for i in range(aux.shape[0])])

#################################################################################################


def save_as_pickle(variable: object, path: str):
    with open(path, 'wb') as handle:
        pickle.dump(variable, handle, protocol=pickle.HIGHEST_PROTOCOL)


def read_pickle(path) -> object:
    with open(path, 'rb') as handle:
        variable = pickle.load(handle)
        return variable


def next_business_day(input_datetime: datetime) -> datetime:
    # Define a list of weekdays that are not considered business days (Saturday and Sunday)
    # 5 represents Saturday, and 6 represents Sunday
    non_business_days = [5, 6]

    # Start with the next day from the input date
    next_day = input_datetime + timedelta(days=1)

    # Check if the next day is a non-business day (Saturday or Sunday)
    while next_day.weekday() in non_business_days:
        next_day += timedelta(days=1)

    return next_day


def get_overview(option, tries=10):
    print("################## Obtaining overview ##################")
    # Download tradingDates and contract_dates
    if option == 1:
        # STOXX
        url = 'https://www.eurex.com/api/v1/overallstatistics/69660'
        print("STOXX")
    else:
        # DAX
        url = 'https://www.eurex.com/api/v1/overallstatistics/70044'
        print("DAX")

    headers = {'Accept': '*/*'}
    params_ov = {'filtertype': 'overview', 'contracttype': 'M'}
    print('Getting data from EUREX...')

    for i in range(tries):
        print(f'try = {i}')
        try:
            response = requests.get(
                url, params=params_ov, headers=headers, timeout=8)

            if not response.ok:
                raise ValueError("Conection failed.")

            tradingDates = pd.to_datetime(
                pd.Series(response.json()['header']['tradingDates']), format="%d-%m-%Y %H:%M")

            # print("Index     Date")
            # for idx,date in enumerate(tradingDates.dt.strftime("%d-%m-%Y")):
            #    print(f" {idx:<6} {date:^3}")

            # date_idxs = get_date_idx(tradingDates.index)
            params_ov['busdate'] = tradingDates[0].strftime("%Y%m%d")
            response = requests.get(
                url, params=params_ov, headers=headers, timeout=8)
            if not response.ok:
                raise ValueError("Conection failed.")

            params_details = {}
            if 'dataRows' in response.json():
                overview_df = pd.DataFrame(response.json()['dataRows'])
                overview_df = overview_df[overview_df.contractType == 'M']
                overview_df.sort_values(
                    'date', ignore_index=True, inplace=True)
                contract_dates = overview_df.loc[:, 'date']
                overview_df.date = pd.to_datetime(
                    overview_df.date, format='%Y%m%d').dt.strftime('%d-%m-%Y')

            else:
                raise ValueError

            params_details['busdate'] = f"{tradingDates[1].strftime('%Y%m%d')}"
            response = requests.get(
                url, params=params_details, headers=headers, timeout=8)
            if not response.ok:
                raise ValueError("Conection failed.")
            break
        except KeyError:
            continue

    return url, headers, tradingDates, contract_dates, overview_df  # type: ignore


def get_contracts(today, url, headers, tradingDates, contract_dates, tries=10):
    # https://en.wikipedia.org/wiki/Offset_(computer_science)#
    offset = 0
    # **tbd: today shall be a working day (Mo - Fr)**

    # busdate = trading_date
    # product_date = contract_date

    expiry = datetime.strptime(contract_dates[0], "%Y%m%d")
    expiry_1 = datetime.strptime(contract_dates[1], "%Y%m%d")

    days_until_expiry = (expiry - today).days

    if days_until_expiry < 0:
        offset = 1

    expiry = datetime.strptime(contract_dates[0 + offset], "%Y%m%d")
    expiry_1 = datetime.strptime(contract_dates[1 + offset], "%Y%m%d")
    days_until_expiry = (expiry - today).days + 1

    print("################## Obtaining contracts ##################")
    params_details = {}
    params_details['filtertype'] = 'detail'
    params_details['contracttype'] = 'M'
    dict_prod_bus = {}
    for productdate_idx in [0, 1]:
        dict_prod_bus[productdate_idx] = {}
        params_details['productdate'] = contract_dates[productdate_idx + offset]
        for busdate_idx in [0, 1]:
            params_details['busdate'] = f"{tradingDates[busdate_idx].strftime('%Y%m%d')}"
            for i in range(tries):
                print(
                    f"busdate_idx = {busdate_idx}   | productdate_idx = {productdate_idx} | try = {i}")
                try:
                    if (busdate_idx == 1) and (productdate_idx == 1):
                        # No need to request data for this case so we skip this iteration
                        break

                    # Request data
                    response = requests.get(
                        url, params=params_details, headers=headers, timeout=10)
                    if not response.ok:
                        raise ValueError("Conection failed.")
                    contractsCall_aux_df = pd.DataFrame(
                        response.json()['dataRowsCall'])
                    contractsPut_aux_df = pd.DataFrame(
                        response.json()['dataRowsPut'])

                    # Create dataframe with the just requested data
                    aux_df = contractsCall_aux_df[['strike', 'openInterest']].merge(
                        contractsPut_aux_df[['strike', 'openInterest']],
                        on="strike",
                        how="left",
                        suffixes=('_CF', '_PF')
                    )
                    dict_prod_bus[productdate_idx][busdate_idx] = aux_df

                    pass
                except KeyError:
                    continue
                break
    return dict_prod_bus, expiry, expiry_1, days_until_expiry


def get_date_idx(list_idxs) -> list:
    while (True):
        date_idxs = [0, 1]
        date_idxs = list(map(int, date_idxs))
        date_idxs = [idx for idx in date_idxs if idx in list_idxs]
        if len(date_idxs) < 1:
            print("No valid indexes passed.")
        else:
            return date_idxs


def get_today():
    today = datetime.now(timezone('Europe/Berlin')).replace(tzinfo=None)
    if today.weekday() > 4:
        # sunday => wd = 6 -> td = +1
        # sartuday => wd = 5 > td +2
        today += timedelta(days=7 - today.weekday())
    return today


def get_euribor_3m() -> float:

    response = requests.get(
        "https://www.euribor-rates.eu/en/current-euribor-rates/2/euribor-rate-3-months/", headers=default_headers)
    euribor_3m_df = pd.read_html(response.text)[
        0].rename(columns={0: "Date", 1: "InterestRate"})
    euribor_row = euribor_3m_df.loc[0]
    print(
        f"Euribor 3M in {euribor_row['Date']} is {euribor_row['InterestRate']}")
    return float(euribor_row["InterestRate"].replace("%", "")) / 100


def get_finazen_price(stock_idx) -> float:
    dict_stock_names = {
        0: 'vdax-new-2m',
        1: 'vstoxx'
    }
    stock = dict_stock_names[stock_idx]
    url = 'https://www.finanzen.net/index/' + stock

    response = requests.get(url, headers=default_headers)
    html = BeautifulSoup(response.text, 'html.parser')
    aux = BeautifulSoup.findAll(
        html, id='snapshot-value-fst-current-0')[0].span.get_text()
    price = re.findall(r'\d{0,},\d{0,}', aux)[0]
    price = float(price.replace(',', '.'))
    print(f"{stock} = {price} PTS")
    return price


def is_calculation_needed(option, days_until_expiry):
    dict_condition = {
        "always": True,
        "days_until_expiry<=5": days_until_expiry < 5
    }
    list_email_send = []
    for email in list_emails:
        if (dict_index_stock[option] == email['index']) and (dict_condition[email['condition']]):
            list_email_send += [email['id']]
    return list_email_send


def parse_eurex(option, today=None):
    """
    option: Option defined by the user. See variables.dict_index_stock.
    today: it should be a date as string using dd/mm/yyyy format. If None, it will be defined by get_today().
    """
    volatility = get_finazen_price(option) / 100

    if not today:
        today = get_today()
    else:
        today = datetime.strptime(today, "%d/%m/%Y")

    print(f"today = {today}")

    # First sdays
    url, headers, tradingDates, contract_dates, overview_df = get_overview(
        option)
    dict_prod_bus, expiry, expiry_1, days_until_expiry = get_contracts(
        today, url, headers, tradingDates, contract_dates)

    list_email_send_selection = is_calculation_needed(
        option, days_until_expiry)

    if len(list_email_send_selection) == 0:
        print(f'Files for {dict_index_stock[option]} were not genereted.')
        return None

    InterestRate = get_euribor_3m()

    response = requests.get(
        "https://www.boerse-stuttgart.de/en/", headers=default_headers)

    stocks = pd.read_html(response.text)

    if option == 0:  # DAX
        stocks_df = stocks[0]
        stock_price = stocks_df.loc[stocks_df['Indices GER'] ==
                                    "L&S DAX", "Price"].values[0]

    else:  # STOXX
        stocks_df = stocks[1]
        stock_price = stocks_df.loc[stocks_df['Indices EU / USA / INT'] ==
                                    "CITI Euro Stoxx 50", "Price"].values[0]

    Spannweite, *_ = get_default_values(option).values()
    span = (Spannweite / 4)
    central_rate = round(stock_price / span) * span

    nbd_today = next_business_day(tradingDates[0])
    nbd_last = tradingDates[0]
    nbd_dict = {'today': nbd_today, 'last': nbd_last}

    print(f'today = {today.strftime("%d/%m/%Y")}')
    print(f"days_until_expiry = {days_until_expiry}")
    print(f"expiry = {expiry}")
    print(f"expiry_1 = {expiry_1}")
    print(f"stock_price = {stock_price}")
    print(f"central_rate = {central_rate}")
    print(f"volatility = {volatility}")
    print(f"InterestRate = {InterestRate}")

    return option, central_rate, volatility, InterestRate, days_until_expiry, nbd_dict, dict_prod_bus, stock_price, expiry, expiry_1, today, list_email_send_selection, contract_dates


def hedge(option, central_rate, volatility, InterestRate, days_until_expiry, dict_prod_bus, stock_price, expiry, expiry_1, today, export_excel=False):

    Spannweite, step, volatility_Laufzeit, contract_value = get_default_values(
        option).values()

    minimum_rate = central_rate - (Spannweite / 2)
    maximum_rate = minimum_rate + Spannweite
    steps = int(Spannweite / step)

    DetailMin = round(stock_price - (Spannweite) / 4)
    DetailMax = DetailMin + (Spannweite / 2)

    if days_until_expiry >= 1:
        delta = 0.5
    else:
        delta = 1

    print(f"minimum_rate = {minimum_rate}")
    print(f"maximum_rate = {maximum_rate}")
    print(f"steps = {steps}")
    print(f"volatility_Laufzeit = {volatility_Laufzeit}")

    # Second sdays
    # Create Basis column that will be used for multiple tables
    Basis = pd.Series((minimum_rate + np.arange(int(steps) + 1) * step)[::-1])

    # Transform series into dataframe to make dataframe creation easier
    Basis_df = Basis.reset_index().rename(columns={0: "Basis"})[["Basis"]]

    # Ueberhaenge_df, sumry_df are created from a copy of the same Dataframe
    # so they have the same index. It means they could be in one  Dataframe but
    #  I decided to keep them separated as they were in the VBA code.

    Ueberhaenge_df = Basis_df.copy()
    sumry_df = Basis_df.copy()

    for productdate_idx in [0, 1]:
        for busdate_idx in [0, 1]:

            if (busdate_idx == 1) and (productdate_idx == 1):
                continue

            # Create dataframe with the just requested data
            aux_df = Basis_df.copy()

            aux_df = aux_df.merge(
                dict_prod_bus[productdate_idx][busdate_idx],
                left_on="Basis",
                right_on="strike",
                how="left"
            )

            if productdate_idx == 0:
                # busdate_idx == 0 -> Front
                # busdate_idx == 1 -> Last
                Ueberhaenge_df[busdate_idx] = aux_df.openInterest_PF - \
                    aux_df.openInterest_CF

                if busdate_idx == 0:
                    sumry_df["openInterest_PF"] = aux_df["openInterest_PF"]
                    sumry_df["openInterest_CF"] = aux_df["openInterest_CF"]

            elif productdate_idx == 1:
                if busdate_idx == 0:
                    Ueberhaenge_df["nextContract"] = (
                        aux_df.openInterest_PF - aux_df.openInterest_CF)
                    sumry_df["nextContract"] = Ueberhaenge_df["nextContract"]

    Ueberhaenge_df.rename(columns={0: "Front"}, inplace=True)
    Ueberhaenge_df.rename(columns={1: "Last"}, inplace=True)

    sumry_df["today"] = Ueberhaenge_df["Front"] * (1 / contract_value) * delta
    sumry_df["last_day"] = Ueberhaenge_df["Last"] * \
        (1 / contract_value) * delta

    Ueberhaenge_df["sum"] = Ueberhaenge_df[[
        "Front", "nextContract"]].sum(axis=1)
    Ueberhaenge_df = Ueberhaenge_df[[
        'Basis', 'sum', "Last", 'Front', "nextContract"]]

    sumry_df["Änderung"] = sumry_df.today - sumry_df.last_day

    if (delta == 1):
        sumry_df['nextContract'] = sumry_df['nextContract'] / 2

    sumryDetail_df = sumry_df[(sumry_df.Basis >= DetailMin) & (
        sumry_df.Basis < (DetailMax + step))]

    # Third sdays
    stepWeite = 10
    Tage = days_until_expiry
    if Tage == 0:
        Tage = 0.5

    Tage_1 = (expiry_1 - today).days

    print(f"Tage = {Tage}")
    print(f"Tage_1 = {Tage_1}")

    rate_count = int(Spannweite / stepWeite) + 1
    HedgeBedarf_kurs = pd.DataFrame(
        maximum_rate - np.arange(rate_count) * stepWeite, columns=["Basis"])

    Hedge_dimensions = rate_count, int(steps + 1)
    HedgeBedarf_values = np.zeros(Hedge_dimensions)
    HedgeBedarf1_values = np.zeros(Hedge_dimensions)

    # Ueberhaenge_df.fillna(0,inplace=True)

    for k in range(Hedge_dimensions[1]):
        Basis_value = Basis[k]
        Kontrakte = Ueberhaenge_df.loc[k, "Front"]
        Kontrakte_1 = Ueberhaenge_df.loc[k,
                                         "nextContract"]  # ASK: Why negative?
        for i in range(Hedge_dimensions[0]):
            rate = maximum_rate - i * stepWeite
            # In python np.log = natural log
            h1 = np.log(rate / Basis_value)
            if option == 0:
                sigma = volatility * ((Tage / volatility_Laufzeit) ** 0.5)
            else:
                sigma = volatility

            sigma_1 = volatility * ((Tage_1 / volatility_Laufzeit) ** 0.5)
            h2 = InterestRate + sigma * sigma / 2
            h2_1 = InterestRate + sigma_1 * sigma_1 / 2
            d1 = (h1 + (h2 * (Tage / 365))) / (sigma * ((Tage / 365) ** 0.5))
            d1_1 = (h1 + (h2_1 * (Tage_1 / 365))) / \
                (sigma_1 * ((Tage_1 / 365) ** 0.5))
            Phi = norm.pdf(d1, 0, 1)
            Phi_1 = norm.pdf(d1_1, 0, 1)
            Gamma = Phi / (rate * (sigma * (Tage / 365) ** 0.5))
            Gamma_1 = Phi_1 / (rate * (sigma_1 * (Tage_1 / 365) ** 0.5))
            HedgeBedarf_values[i, k] = Gamma * Kontrakte / contract_value
            HedgeBedarf1_values[i, k] = Gamma_1 * Kontrakte_1 / contract_value

            if rate == 17000:
                continue

    HedgeSum = HedgeBedarf_values.sum(axis=1) / 2
    HedgeSum_1 = HedgeBedarf1_values.sum(axis=1) / 2

    HedgeBedarf_df = pd.DataFrame(data=HedgeBedarf_values, columns=Basis)
    HedgeSum_df = pd.DataFrame(HedgeSum, columns=["HedgeSum"])
    HedgeBedarf_df = pd.concat(
        [HedgeBedarf_kurs, HedgeSum_df, HedgeBedarf_df], axis=1)

    HedgeBedarf1_df = pd.DataFrame(data=HedgeBedarf1_values, columns=Basis)
    HedgeSum1_df = pd.DataFrame(HedgeSum_1, columns=["HedgeSum"])
    HedgeBedarf1_df = pd.concat(
        [HedgeBedarf_kurs, HedgeSum1_df, HedgeBedarf1_df], axis=1)

    # Fourth Sdays
    info_df = pd.DataFrame({
        'central_rate': [central_rate],
        'Spannweite': [Spannweite],
        'stepWeite': [stepWeite],
        'step': [step],
        'option': dict_index_stock[option],
        'expiry': expiry,
        'expiry_1': expiry_1,
        'today': today,
        'days_until_expiry': days_until_expiry,
        'stock_price': stock_price,
        'minimum_rate': minimum_rate,
        'maximum_rate': maximum_rate,
        'steps': steps,
        'DetailMin': DetailMin,
        'DetailMax': DetailMax,
        'delta': delta,
        'volatility': volatility
    }).T.reset_index().rename(columns={0: "Value", 'index': "Info"})

    if export_excel:
        list_excel_files = [
            os.path.join(current_results_path,
                         f"{dict_index_stock[option]}.xlsx"),
            os.path.join(
                old_results_path, f"{dict_index_stock[option]}_{today.strftime('%Y_%m_%d')}.xlsx")
        ]

        for excel_file in list_excel_files:
            with pd.ExcelWriter(excel_file, datetime_format="DD/MM/YYYY") as writer:
                info_df.to_excel(writer, sheet_name='infos', index=False)
                Ueberhaenge_df.to_excel(
                    writer, sheet_name='Ueberhaenge', index=False)
                sumry_df.to_excel(writer, sheet_name='sumry', index=False)
                sumryDetail_df.to_excel(
                    writer, sheet_name='sumryDetail', index=False)
                HedgeBedarf_df.to_excel(
                    writer, sheet_name='HedgeBedarf', index=False)
                HedgeBedarf1_df.to_excel(
                    writer, sheet_name='HedgeBedarf+01', index=False)

        print("Excel files have been exported.")
    return (
        sumry_df,  # .dropna(ignore_index = True),
        HedgeBedarf_df,  # .dropna(ignore_index = True),
        HedgeBedarf1_df,  # .dropna(ignore_index = True),
        Ueberhaenge_df,  # .dropna(ignore_index = True),
        delta)


def parse_excel(option: int, excel_path: str):
    """
    Parse excel file and return the necessary variables to perform the hedge calculation and compare the results with the ones from the excel file.
    """

    dict_option_prefix = {
        0: "",
        1: "STOXX_"
    }

    wb = load_workbook(excel_path, data_only=True, keep_vba=True)
    Ueberhaenge_sheet = wb['Ueberhaenge']
    today = Ueberhaenge_sheet.cell(*coordinate_to_tuple("G5")).value
    expiry = Ueberhaenge_sheet.cell(*coordinate_to_tuple("G3")).value
    expiry_1 = Ueberhaenge_sheet.cell(*coordinate_to_tuple("R6")).value
    stock_price = Ueberhaenge_sheet.cell(*coordinate_to_tuple("X4")).value
    InterestRate = Ueberhaenge_sheet.cell(*coordinate_to_tuple("N3")).value
    central_rate = Ueberhaenge_sheet.cell(*coordinate_to_tuple("C3")).value
    volatility = Ueberhaenge_sheet.cell(*coordinate_to_tuple("N4")).value

    contract_dates = [expiry, expiry_1]

    sumry_sheet = wb[f'{dict_option_prefix[option]}sumry']
    nbd_today = sumry_sheet.cell(*coordinate_to_tuple("C9")).value

    nbd_last = sumry_sheet.cell(*coordinate_to_tuple("C9")).value
    nbd_last = next_business_day(today)  # type: ignore
    nbd_dict = {'today': nbd_today, 'last': nbd_last}

    days_until_expiry = (expiry - today).days  # type: ignore

    list_email_send_selection = is_calculation_needed(
        option, days_until_expiry)

    dict_prod_bus = {}
    for productdate_idx in [0, 1]:
        dict_prod_bus[productdate_idx] = {}
        for busdate_idx in [0, 1]:
            print(
                f"busdate_idx = {busdate_idx}   | productdate_idx = {productdate_idx}")
            if (busdate_idx == 1) and (productdate_idx == 1):
                # No need to request data for this case so we skip this iteration
                break

            if (busdate_idx == 0) and (productdate_idx == 0):
                contractsCall_aux_df = pd.read_excel(
                    excel_path, sheet_name=f"{dict_option_prefix[option]}Call_Front")
                contractsPut_aux_df = pd.read_excel(
                    excel_path, sheet_name=f"{dict_option_prefix[option]}Put_Front")

            if (busdate_idx == 1) and (productdate_idx == 0):
                contractsCall_aux_df = pd.read_excel(
                    excel_path, sheet_name=f"{dict_option_prefix[option]}CallFront-1")
                contractsPut_aux_df = pd.read_excel(
                    excel_path, sheet_name=f"{dict_option_prefix[option]}PutFront-1")

            if (busdate_idx == 0) and (productdate_idx == 1):
                contractsCall_aux_df = pd.read_excel(
                    excel_path, sheet_name=f"{dict_option_prefix[option]}Put+01")
                contractsPut_aux_df = pd.read_excel(
                    excel_path, sheet_name=f"{dict_option_prefix[option]}Call+01")

            aux_df = contractsCall_aux_df[['strike', 'openInterest']].merge(  # type: ignore
                contractsPut_aux_df[  # type: ignore
                    ['strike', 'openInterest']],
                on="strike",
                how="left",
                suffixes=('_CF', '_PF')
            )
            dict_prod_bus[productdate_idx][busdate_idx] = aux_df

    print(f'today = {today.strftime("%d/%m/%Y")}')  # type: ignore
    print(f"days_until_expiry = {days_until_expiry}")
    print(f"expiry = {expiry}")
    print(f"expiry_1 = {expiry_1}")
    print(f"stock_price = {stock_price}")
    print(f"central_rate = {central_rate}")
    print(f"volatility = {volatility}")
    print(f"InterestRate = {InterestRate}")

    return option, central_rate, volatility, InterestRate, days_until_expiry, nbd_dict, dict_prod_bus, stock_price, expiry, expiry_1, today, list_email_send_selection, contract_dates


def generate_pdfs(option, sumry_df, HedgeBedarf_df, HedgeBedarf1_df, stock_price, today, nbd_dict, days_until_expiry, delta, expiry, expiry_1):
    """
    Generate PDFs for the report formats: basic, complete and detailed.
    """

    # Drop rows with NaN values
    sumry_df.dropna(ignore_index=True, inplace=True)
    HedgeBedarf_df.dropna(ignore_index=True, inplace=True)
    HedgeBedarf1_df.dropna(ignore_index=True, inplace=True)

    # Fifth sdays
    min_Kontrakte = 5000
    prozentual = 0.2

    dict_expiry_sufix = {
        "detailed": "_expiry",
        "complete": "",
        "basic": "_basic"
    }

    # True if the expiry is close
    is_close_expiry = days_until_expiry < 5

    # Report format for which the images will be generated => images for complete report format are always generated
    report_formats = ['complete']

    # PDF formats => complete is always generated
    pdf_formats = ['complete']

    if option == 0:
        # Also generate images basic report format
        report_formats.append('basic')

        # Also generate basic report pdf format
        pdf_formats.append('basic')

    if is_close_expiry:
        # Also generate images detailed report format. This is only done if the expiry is close
        # These images have to be the last ones to be generated because the dataframes are clipped
        # around the stock price. This requirment can be removed if we make a copy of the dataframes.
        report_formats.append('detailed')

    # Generate images for each report format
    for report_format in report_formats:

        if report_format == 'detailed':
            ################## Select 10 values above and below the stock price ##################
            indexes = sumry_df.loc[(sumry_df.Basis >= stock_price)].tail(10).index.to_list(
            ) + sumry_df.loc[(sumry_df.Basis < stock_price)].head(10).index.to_list()
            sumry_df = sumry_df.loc[indexes].reset_index(drop=True)

            indexes = HedgeBedarf_df.loc[(HedgeBedarf_df.Basis >= sumry_df.Basis.min()) & (
                HedgeBedarf_df.Basis <= sumry_df.Basis.max())].index

            HedgeBedarf_df = HedgeBedarf_df.loc[indexes].reset_index(drop=True)
            HedgeBedarf1_df = HedgeBedarf1_df.loc[indexes].reset_index(
                drop=True)

            ######################################################################################

        # Closest index to the stock price
        idx_closest = (HedgeBedarf_df.Basis - stock_price).abs().idxmin()

        # Closest Basis to the stock price
        closest_Basis = HedgeBedarf_df.loc[idx_closest, "Basis"]

        # Define w
        if report_format == 'basic':
            line_to_point = HedgeBedarf_df.HedgeSum
            line_to_point_legend_name = expiry.strftime("%Y-%m")
        else:
            # report_format == 'complete' or report_format == 'detailed
            hedge_sum = HedgeBedarf_df.HedgeSum + HedgeBedarf1_df.HedgeSum
            line_to_point = hedge_sum
            line_to_point_legend_name = expiry.strftime(
                "%Y-%m") + " + " + expiry_1.strftime("%Y-%m")

        # Where to point the arrow
        value_to_point = line_to_point.loc[idx_closest]

        ################################## Hilights for Basis column ####################################

        sumry_df["basis_color"] = "lavender"
        if report_format == 'complete':
            # (CF > min_Kontrakte) AND (PF > min_Kontrakte) AND (ABS(CF - PF)  < MIN(PF,CF)*)
            mask_highlights = (
                (sumry_df.openInterest_CF > min_Kontrakte) &
                (sumry_df.openInterest_PF > min_Kontrakte) &
                (
                    (sumry_df.openInterest_CF - sumry_df.openInterest_PF).abs() <
                    (sumry_df[['openInterest_PF', "openInterest_CF"]].min(axis=1) * prozentual))
            )

            # Highlight the Basis
            sumry_df.loc[
                mask_highlights, "basis_color"] = "yellow"
        col_basis_color = sumry_df.basis_color.to_numpy()
        #################################################################################################

        ######################### Gradient of red and blue for Anderung column ##########################
        aux = sumry_df.Änderung.reset_index()
        aux['r'] = aux['g'] = aux['b'] = 255
        aux.loc[aux.Änderung < 0, 'g'] = aux.loc[aux.Änderung < 0,
                                                 'b'] = 255 - 150 * aux.Änderung / (aux.Änderung.min())
        aux.loc[aux.Änderung > 0, 'g'] = aux.loc[aux.Änderung > 0,
                                                 'r'] = 255 - 150 * aux.Änderung / (aux.Änderung.max())
        aux = aux.astype(float)
        rb_shades = np.array(
            [f"rgb({aux.loc[i,'r']},{aux.loc[i,'g']},{aux.loc[i,'b']})" for i in range(aux.shape[0])])
        col_anderung_color = rb_shades
        #################################################################################################

        ############################# Colors for today and last_day columns #############################
        col_today_color = posneg_binary_color(
            sumry_df.today, "rgb(0, 204, 204)", "rgb(77, 166, 255)")
        col_last_day_color = posneg_binary_color(
            sumry_df.last_day, "rgb(0, 204, 204)", "rgb(77, 166, 255)")
        #################################################################################################

        ###################### Gradient of green and blue for Put anc Call columns ######################
        col_put_color = np.array(n_colors('rgb(214, 245, 214)', 'rgb(40, 164, 40)',
                                          20, colortype='rgb'))[scale_col_range(sumry_df.openInterest_PF, 19)]

        col_call_color = np.array(n_colors('rgb(204, 224, 255)', 'rgb(0, 90, 179)',
                                           20, colortype='rgb'))[scale_col_range(sumry_df.openInterest_CF, 19)]
        #################################################################################################

        num_rows = sumry_df.shape[0]  # Not including the header
        header_height = 20  # Minimum height of the header
        height = 1050
        row_height = (height - header_height) // (num_rows)

        header_height += height - header_height - row_height * \
            num_rows  # Adjust the header height to fit the image height

        values_body = [
            "<b>" + (sumry_df.Basis +
                     1e-6).round(0).astype(int).astype(str) + "</b>",
            (sumry_df.Änderung + 1e-6).round(0).astype(int),
            (sumry_df.today + 1e-6).round(0).astype(int),
            (sumry_df.last_day + 1e-6).round(0).astype(int),
        ]

        values_header = bold(["Basis", "Änderung", nbd_dict['today'].strftime(  # type: ignore
            "%d/%m/%y"), nbd_dict['last'].strftime("%d/%m/%y")])

        if report_format == 'complete':
            values_body += [(sumry_df.openInterest_PF),
                            (sumry_df.openInterest_CF)]
            values_header += bold(["Put", "Call"])  # type: ignore

        font_size = int(440 / num_rows)

        fig = go.Figure(
            data=[
                go.Table(

                    # Define some paremeters for the header
                    header=dict(
                        # Names of the columns
                        values=values_header,

                        # Header style
                        fill_color='paleturquoise',
                        align='center',
                        font={'size': 9},
                        height=header_height,
                    ),

                    cells=dict(

                        align='center',
                        height=row_height,
                        font={
                            'size': [font_size - 3] + [font_size for i in range(len(values_header) - 1)]},

                        # Values of the table
                        values=values_body,

                        # Colors of the columns
                        fill_color=[
                            col_basis_color,
                            col_anderung_color,
                            col_today_color,
                            col_last_day_color,
                            col_put_color,
                            col_call_color
                        ],
                    )
                )
            ],
        )
        fig.update_layout(
            height=1050, width=550,
            margin=dict(l=0, r=0, b=0.0, t=0)
        )

        fig.write_image(os.path.join(
            temp_results_path, f'image_table{dict_expiry_sufix[report_format]}.svg'), scale=1)

        # Length x axis
        x_axis_length = line_to_point.max() - line_to_point.min()
        data = [

            # This plot the 0 line
            go.Scatter(
                x=[0, 0],
                y=[HedgeBedarf_df.Basis.min(), HedgeBedarf_df.Basis.max()],
                mode="lines",
                marker_color="orange",
                showlegend=False
            ),


            # This plot the arrow pointing to line_to_point
            go.Scatter(
                x=[value_to_point + x_axis_length / 5,
                    value_to_point + x_axis_length / 50],
                y=[closest_Basis, closest_Basis],
                marker=dict(size=20, symbol="arrow-bar-up",
                            angleref="previous"),
                marker_color="red",
                showlegend=False
            ),


            go.Scatter(
                x=line_to_point,
                y=HedgeBedarf_df.Basis,
                mode="lines",
                name=line_to_point_legend_name,
                marker_color="blue"
            )

        ]

        # Define range for x axis
        min_x_list = [line_to_point.min()]
        max_x_list = [line_to_point.max()]

        # If the report is complete, add the second line
        if report_format == 'complete':
            min_x_list.append(HedgeBedarf1_df.HedgeSum.min())
            max_x_list.append(HedgeBedarf1_df.HedgeSum.max())
            data += [
                go.Scatter(
                    x=HedgeBedarf1_df.HedgeSum,
                    y=HedgeBedarf_df.Basis,
                    mode="lines",
                    name=expiry_1.strftime("%Y-%m"),
                    marker_color="rgb(255,0,255)"
                ),
            ]

        ##################################################################################################

        # Margin to the x axis
        dx = 0.2 * (hedge_sum.max() - hedge_sum.min())  # type: ignore

        fig = go.Figure(data=data)
        fig.update_layout(
            margin=dict(l=0, r=0, b=row_height // 2 + 300 //
                        row_height, t=header_height + row_height // 2),
            # Set limits in the x and y axis
            yaxis_range=[HedgeBedarf_df.Basis.min(),
                         HedgeBedarf_df.Basis.max()],
            xaxis_range=[min(min_x_list) - dx, max(max_x_list) + dx],
            yaxis=dict(
                tickmode='array',
                tickvals=sumry_df.Basis,  # HedgeBedarf_df.Basis[mask],
                ticktext=sumry_df.Basis,  # HedgeBedarf_df.Basis[mask]
            ),

            # Remove margins
            paper_bgcolor="white",

            # Define width and height of the image
            width=380, height=1050,
            template="seaborn",

            # Legend parameters
            legend=dict(
                yanchor="top",
                y=1,  # - (row_height/height),
                xanchor="right",
                x=1,
                font=dict(
                    size=8
                ),
                # legend in the vertical
                orientation="v",
                bgcolor='rgba(0,0,0,0)'
            ),
        )

        fig.write_image(os.path.join(
            temp_results_path, f'image_graph{dict_expiry_sufix[report_format]}.svg'), scale=1)

        fig = go.Figure()

        if report_format == 'complete':
            fig.add_trace(
                trace=go.Bar(name='Put', y=sumry_df.Basis, x=-sumry_df.openInterest_PF,
                             orientation='h', marker_color='rgb(40, 164, 40)'),
            )

            fig.add_trace(
                trace=go.Bar(name='Call', y=sumry_df.Basis, x=sumry_df.openInterest_CF,
                             orientation='h', marker_color='rgb(0, 90, 179)'),
            )

        fig.update_layout(
            # barmode='stack',
            barmode='overlay',
            margin=dict(l=0, r=0, b=0, t=header_height),
            # Set limits in the x and y axis
            # yaxis_range= [HedgeBedarf_df.Basis.min(), HedgeBedarf_df.Basis.max()],
            # xaxis_range= [hedge_sum.min() - dx, hedge_sum.max() + dx],

            # Define width and height of the image
            width=80, height=1050,
            template="seaborn",
            showlegend=False,
            bargap=0.5,
            plot_bgcolor='rgba(0, 0, 0, 0)',
            paper_bgcolor='rgba(0, 0, 0, 0)',
        )

        fig.update_xaxes(visible=False)
        fig.update_yaxes(visible=False)

        fig.write_image(os.path.join(
            temp_results_path, f'image_bar{dict_expiry_sufix[report_format]}.svg'), scale=1)

    dict_option_colors = {
        0: {"$HEADER_COLOR$": "", "$FUTURE_COLOR$": "background-color: rgb(0, 174, 255);", "$FONT_COLOR$": "color: black;"},
        1: {"$HEADER_COLOR$": "background-color: rgb(12, 89, 177);", "$FUTURE_COLOR$": "", "$FONT_COLOR$": "color: white;"},
    }

    with open(src_css) as file:
        css_file = file.read()

    for key, value in dict_option_colors[option].items():
        css_file = css_file.replace(key, str(value))

    with open(result_css, 'w') as file:
        file.write(css_file)

    dict_html_sections = {
        "complete_report": "COMPLETE_REPORT",
        "detailed_report": "DETAIL",
        "basic_report": "BASIC",
        "after_report": "AFTER_REPORT"
    }

    dict_title = {
        0: "OpenInterest und HedgeBedarf",
        1: "STOXX 50 OpenInterest und HedgeBedarf",
    }

    # Replace specific characters in the template by values
    dict_raplace = {
        "$DATE$": today.strftime("%d/%m/%Y"),
        "$PUT_SUM$": int(sumry_df.openInterest_PF.sum()),
        "$CALL_SUM$": int(sumry_df.openInterest_CF.sum()),
        "$TBF$": days_until_expiry,
        "$DELTA$": str(delta).replace(".", ","),
        "$FRONT_DATE$": expiry.strftime("%Y-%m"),
        "$TITLE$": dict_title[option],
        "$HEADER_COLOR$": "background-color: rgb(12, 89, 177)"
    }

    with open(src_html) as file:
        template_original = file.read()

    template_replaced_values = template_original

    for key, value in dict_raplace.items():
        template_replaced_values = template_replaced_values.replace(
            key, str(value))

    for pdf_format in pdf_formats:
        template = template_replaced_values
        if pdf_format == "basic":
            maintain_html_list = ["basic_report", "after_report"]

        elif pdf_format == "complete":
            maintain_html_list = ["complete_report", "after_report"]

        if is_close_expiry:
            maintain_html_list.append("detailed_report")  # type: ignore

        for html_section, html_section_name in dict_html_sections.items():
            if html_section in maintain_html_list:  # type: ignore
                template = template.replace(f"$BEGIN_{html_section_name}$", "").replace(
                    f"$END_{html_section_name}$", "")
            else:
                subst = ""
                regex = f"\$BEGIN_{html_section_name}\$.*\$END_{html_section_name}\$"
                template = re.sub(regex, subst, template, 0,
                                  re.MULTILINE | re.DOTALL)

        result_html = os.path.join(temp_results_path, "output.html")

        # Export html file
        with open(result_html, 'w') as file:
            file.write(template)

        pdf_suffix = {"complete": "", "basic": "_basic"}[pdf_format]
        file_path = os.path.join(
            current_results_path, f"{dict_index_stock[option]}_{today.strftime('%d_%m_%Y')}{pdf_suffix}.pdf")

        converter.convert(
            "file://" + os.path.join(os.getcwd(), result_html), file_path)

        print("PDF has been generated.")
